import calendar as cal_module
import json
from collections import defaultdict
from datetime import datetime, timedelta

import nepali_datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.cache import never_cache

from daily_coverage.models import ChemistCoverage, DailyCoverage, StockistCoverage
from doctor_employee_relation.models import DoctorEmployeeRelation
from tour_plans.models import TourPlan

SUPER_CORE_MAX = 25
CORE_MAX = 75

VISIT_TARGETS = {"super_core": 4, "core": 2, "vip": 1}
CATEGORY_LABELS = {"super_core": "Super Core", "core": "Core", "vip": "VIP"}

# The yearly report runs on Bikram Sambat years (Baishakh … Chaitra)
BS_MONTH_NAMES = ["Baishakh", "Jestha", "Asar", "Shrawan", "Bhadau", "Aswin",
                  "Kartik", "Mangsir", "Poush", "Magh", "Falgun", "Chaitra"]
BS_MONTH_ABBR = ["Bai", "Jes", "Asa", "Shr", "Bha", "Asw",
                 "Kar", "Man", "Pou", "Mag", "Fal", "Cha"]


def _get_employee(request):
    """Resolve which employee's report to show.

    Visibility follows the position hierarchy (users.User.TYPE_RANK): a user
    sees their own reports plus those of anyone at a strictly lower position;
    superusers see everyone. Requesting anyone outside that set 404s.
    """
    viewer = request.user
    viewable = viewer.viewable_report_users()
    try:
        employee_id = int(request.GET.get("employee_id") or "")
    except ValueError:
        employee_id = None
    employee = get_object_or_404(viewable, pk=employee_id) if employee_id else viewer
    can_view_others = viewable.exclude(pk=viewer.pk).exists()
    all_employees = viewable if can_view_others else []
    return employee, all_employees, employee_id, can_view_others


def _doctor_category(msl):
    if msl and 1 <= msl <= SUPER_CORE_MAX:
        return "super_core"
    if msl and SUPER_CORE_MAX < msl <= CORE_MAX:
        return "core"
    return "vip"


def _build_yearly_data(employee, year):
    """
    Yearly report data for a **BS year**: rows (one per approved
    doctor-employee relation, ordered by MSL then doctor name, with
    month_visits as 12 BS-month day-lists) plus per-BS-month visit
    frequency by doctor category for the chart.
    """
    try:
        start_ad = nepali_datetime.date(year, 1, 1).to_datetime_date()
        end_ad = nepali_datetime.date(year + 1, 1, 1).to_datetime_date() - timedelta(days=1)
    except ValueError:  # outside nepali-datetime's supported range
        start_ad = end_ad = None

    relations = (
        DoctorEmployeeRelation.objects
        .filter(employee=employee, status=DoctorEmployeeRelation.Status.APPROVED)
        .select_related("doctor", "doctor__hospital")
        .order_by("msl_number", "doctor__name")
    )

    all_covs = []
    if start_ad:
        all_covs = list(
            DailyCoverage.objects
            .filter(created_by=employee, report_date__range=(start_ad, end_ad))
            .order_by("report_date")
        )
    for c in all_covs:
        c.bs = nepali_datetime.date.from_datetime_date(c.report_date)

    # Chart: visits per BS month split by doctor category (msl over *all*
    # relations, mirroring the monthly report's classification)
    msl_map = {
        rel.doctor_id: rel.msl_number
        for rel in DoctorEmployeeRelation.objects.filter(employee=employee)
    }
    freq = {m: {"super_core": 0, "core": 0, "vip": 0} for m in range(1, 13)}
    for c in all_covs:
        freq[c.bs.month][_doctor_category(msl_map.get(c.doctor_id))] += 1

    cov_by_doctor = defaultdict(list)
    for c in all_covs:
        cov_by_doctor[c.doctor_id].append(c)

    rows = []
    for rel in relations:
        doc = rel.doctor
        covs = cov_by_doctor.get(doc.id, [])

        month_visits = []
        for m in range(1, 13):
            days = sorted({c.bs.day for c in covs if c.bs.month == m})
            month_visits.append(days)

        rows.append({
            "msl":         rel.msl_number or "-",
            "doctor":      doc,
            "hospitals":   [doc.hospital.name],
            "month_visits": month_visits,
            "total_calls": len(covs),
        })

    months = list(range(1, 13))
    return {
        "rows": rows,
        "freq": freq,
        "total_super_core": sum(freq[m]["super_core"] for m in months),
        "total_core":       sum(freq[m]["core"]       for m in months),
        "total_vip":        sum(freq[m]["vip"]        for m in months),
    }


# ── Daily Activity Report ────────────────────────────────────────────────────

@login_required
@never_cache
def daily_activity_report(request):
    employee, all_employees, selected_employee_id, can_view_others = _get_employee(request)
    date_str = request.GET.get("date", "")

    report_date = None
    if date_str:
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    coverages = []
    planned_areas = []
    worked_areas = []
    work_day_label = ""
    chemist_coverages = []
    stockist_coverages = []

    if report_date:
        qs = (
            DailyCoverage.objects
            .filter(created_by=employee, report_date=report_date)
            .select_related("doctor", "doctor__hospital", "actual_working_place")
            .order_by("call_time")
        )
        coverages = list(qs)

        doctor_ids = [c.doctor_id for c in coverages]
        msl_map = {
            rel.doctor_id: rel.msl_number
            for rel in DoctorEmployeeRelation.objects.filter(
                employee=employee, doctor_id__in=doctor_ids,
            )
        }
        for c in coverages:
            c.msl = msl_map.get(c.doctor_id) or "-"

        planned_areas = list(
            TourPlan.objects
            .filter(created_by=employee, plan_date=report_date, status=TourPlan.Status.APPROVED)
            .select_related("area")
            .values_list("area__name", flat=True)
        )
        worked_areas = list(dict.fromkeys(c.actual_working_place.name for c in coverages))
        work_day_label = coverages[0].get_work_day_display() if coverages else ""

        chemist_coverages = list(
            ChemistCoverage.objects
            .filter(created_by=employee, report_date=report_date)
            .select_related("area")
            .order_by("call_time")
        )
        stockist_coverages = list(
            StockistCoverage.objects
            .filter(created_by=employee, report_date=report_date)
            .select_related("area")
            .order_by("call_time")
        )

    return render(request, "reports/daily_activity_report.html", {
        "date_str": date_str,
        "report_date": report_date,
        "employee": employee,
        "all_employees": all_employees,
        "selected_employee_id": selected_employee_id,
        "coverages": coverages,
        "planned_areas": planned_areas,
        "worked_areas": worked_areas,
        "work_day_label": work_day_label,
        "chemist_coverages": chemist_coverages,
        "stockist_coverages": stockist_coverages,
        "can_view_others": can_view_others,
    })


# ── Monthly Activity Report ──────────────────────────────────────────────────

def _parse_month(month_str):
    """Return (year, month, normalized "YYYY-MM"); defaults to the current month."""
    if month_str:
        try:
            dt = datetime.strptime(month_str, "%Y-%m")
            return dt.year, dt.month, month_str
        except ValueError:
            pass
    today = datetime.today()
    return today.year, today.month, f"{today.year}-{today.month:02d}"


def _parse_bs_month(request):
    """Resolve the requested **BS** month from `year` + `bs_month` selects,
    defaulting to the current BS month. Returns
    (bs_year, bs_month, "YYYY-MM", start_ad, end_ad) where the AD span covers
    exactly that BS month."""
    today_bs = nepali_datetime.date.from_datetime_date(datetime.today().date())
    try:
        year = int(request.GET.get("year", ""))
        month = int(request.GET.get("bs_month", ""))
        if not 1 <= month <= 12:
            raise ValueError
        start = nepali_datetime.date(year, month, 1)
    except (ValueError, TypeError):
        year, month = today_bs.year, today_bs.month
        start = nepali_datetime.date(year, month, 1)

    nxt = (nepali_datetime.date(year + 1, 1, 1) if month == 12
           else nepali_datetime.date(year, month + 1, 1))
    start_ad = start.to_datetime_date()
    end_ad = nxt.to_datetime_date() - timedelta(days=1)
    return year, month, f"{year}-{month:02d}", start_ad, end_ad


def _build_monthly_data(employee, start_ad, end_ad):
    """Data shared by the Monthly Activity report page and its Excel export.
    Scoped to a BS month's AD span; days and frequencies are keyed by BS day."""
    num_days = (end_ad - start_ad).days + 1

    msl_map = {
        rel.doctor_id: rel.msl_number
        for rel in DoctorEmployeeRelation.objects.filter(employee=employee)
    }

    coverages = list(
        DailyCoverage.objects
        .filter(created_by=employee, report_date__range=(start_ad, end_ad))
        .select_related("doctor", "actual_working_place")
        .order_by("report_date", "call_time")
    )
    for c in coverages:
        c.msl = msl_map.get(c.doctor_id)
        c.category = _doctor_category(c.msl)
        c.bs = nepali_datetime.date.from_datetime_date(c.report_date)

    freq = {d: {"super_core": 0, "core": 0, "vip": 0} for d in range(1, num_days + 1)}
    for c in coverages:
        freq[c.bs.day][c.category] += 1

    days = list(range(1, num_days + 1))

    all_specs = sorted({c.doctor.specialization for c in coverages if c.doctor.specialization})

    tour_by_date = defaultdict(list)
    for tp in TourPlan.objects.filter(
        created_by=employee, plan_date__range=(start_ad, end_ad),
        status=TourPlan.Status.APPROVED,
    ).select_related("area").order_by("plan_date"):
        tour_by_date[tp.plan_date].append(tp.area.name)

    cov_by_date = defaultdict(list)
    for c in coverages:
        cov_by_date[c.report_date].append(c)

    rows = []
    for d in sorted(cov_by_date):
        day_covs = cov_by_date[d]
        spec_counts = defaultdict(int)
        for c in day_covs:
            if c.doctor.specialization:
                spec_counts[c.doctor.specialization] += 1
        d_bs = nepali_datetime.date.from_datetime_date(d)
        rows.append({
            "date":        d,
            "date_bs":     f"{d_bs.day} {BS_MONTH_ABBR[d_bs.month - 1]} {d_bs.year}",
            "tour_areas":  tour_by_date.get(d, []),
            "actual_areas": list(dict.fromkeys(c.actual_working_place.name for c in day_covs)),
            "spec_values": [spec_counts.get(s, 0) for s in all_specs],
            "total_drs":   len(day_covs),
        })

    return {
        "days":             days,
        "freq":             freq,
        "all_specs":        all_specs,
        "rows":             rows,
        "total_super_core": sum(freq[d]["super_core"] for d in days),
        "total_core":       sum(freq[d]["core"]       for d in days),
        "total_vip":        sum(freq[d]["vip"]        for d in days),
    }


@login_required
@never_cache
def monthly_activity_report(request):
    employee, all_employees, selected_employee_id, can_view_others = _get_employee(request)
    year, month, month_str, start_ad, end_ad = _parse_bs_month(request)

    data = _build_monthly_data(employee, start_ad, end_ad)

    chart_data = json.dumps({
        "labels":     data["days"],
        "super_core": [data["freq"][d]["super_core"] for d in data["days"]],
        "core":       [data["freq"][d]["core"]       for d in data["days"]],
        "vip":        [data["freq"][d]["vip"]        for d in data["days"]],
    })

    return render(request, "reports/monthly_activity_report.html", {
        "month_str":           month_str,
        "year":                year,
        "month":               month,
        "month_name":          BS_MONTH_NAMES[month - 1],
        "year_choices":        _year_choices(),
        "bs_months":           BS_MONTH_NAMES,
        "employee":            employee,
        "all_employees":       all_employees,
        "selected_employee_id": selected_employee_id,
        "can_view_others":     can_view_others,
        "chart_data":          chart_data,
        "total_super_core":    data["total_super_core"],
        "total_core":          data["total_core"],
        "total_vip":           data["total_vip"],
        "rows":                data["rows"],
        "all_specs":           data["all_specs"],
    })


@login_required
def monthly_activity_report_excel(request):
    employee, _, _, _ = _get_employee(request)
    year, month, _, start_ad, end_ad = _parse_bs_month(request)
    data = _build_monthly_data(employee, start_ad, end_ad)

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="4A90C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _style_header(sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    # Sheet 1 — the "List of data" table
    ws = wb.active
    ws.title = "Daily Calls"
    ws.append(
        ["SN", "Date", "Tour Plan", "Actual Work Plan"]
        + data["all_specs"]
        + ["Total DRS Calls"]
    )
    _style_header(ws)
    for i, row in enumerate(data["rows"], start=1):
        ws.append([
            i,
            row["date_bs"],
            " / ".join(row["tour_areas"]) or "-",
            " / ".join(row["actual_areas"]) or "-",
            *row["spec_values"],
            row["total_drs"],
        ])
    col_widths = [6, 14, 22, 22] + [14] * len(data["all_specs"]) + [14]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Sheet 2 — MSL frequency per day (the chart's data)
    ws2 = wb.create_sheet("MSL Frequency")
    ws2.append(["Day", "Super Core", "Core", "VIP", "Total"])
    _style_header(ws2)
    for d in data["days"]:
        day_freq = data["freq"][d]
        ws2.append([
            d,
            day_freq["super_core"],
            day_freq["core"],
            day_freq["vip"],
            day_freq["super_core"] + day_freq["core"] + day_freq["vip"],
        ])
    total_row = [
        "Total",
        data["total_super_core"],
        data["total_core"],
        data["total_vip"],
        data["total_super_core"] + data["total_core"] + data["total_vip"],
    ]
    ws2.append(total_row)
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
    for col_idx, width in enumerate([10, 12, 10, 10, 10], start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    emp_name = employee.get_full_name() or employee.username
    response["Content-Disposition"] = (
        f'attachment; filename="monthly_report_{emp_name}_{year}-{month:02d}.xlsx"'
    )
    wb.save(response)
    return response


# ── Yearly Activity Report ───────────────────────────────────────────────────

def _current_bs_year():
    return nepali_datetime.date.from_datetime_date(datetime.today().date()).year


def _year_choices():
    current = _current_bs_year()
    return list(range(current - 4, current + 2))


@login_required
@never_cache
def yearly_activity_report(request):
    employee, all_employees, selected_employee_id, can_view_others = _get_employee(request)
    year_str = request.GET.get("year", str(_current_bs_year()))
    try:
        year = int(year_str)
    except ValueError:
        year = _current_bs_year()

    data = _build_yearly_data(employee, year)

    chart_data = json.dumps({
        "labels":     BS_MONTH_NAMES,
        "super_core": [data["freq"][m]["super_core"] for m in range(1, 13)],
        "core":       [data["freq"][m]["core"]       for m in range(1, 13)],
        "vip":        [data["freq"][m]["vip"]        for m in range(1, 13)],
    })

    return render(request, "reports/yearly_activity_report.html", {
        "year":                year,
        "year_choices":        _year_choices(),
        "employee":            employee,
        "all_employees":       all_employees,
        "selected_employee_id": selected_employee_id,
        "can_view_others":     can_view_others,
        "rows":                data["rows"],
        "month_abbr":          BS_MONTH_ABBR,
        "chart_data":          chart_data,
        "total_super_core":    data["total_super_core"],
        "total_core":          data["total_core"],
        "total_vip":           data["total_vip"],
    })


@login_required
def yearly_activity_report_excel(request):
    employee, _, _, _ = _get_employee(request)
    year_str = request.GET.get("year", str(_current_bs_year()))
    try:
        year = int(year_str)
    except ValueError:
        year = _current_bs_year()

    rows = _build_yearly_data(employee, year)["rows"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Yearly Report {year}"

    header_fill = PatternFill("solid", fgColor="4A90C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = (
        ["MSL No.", "Doctor's Name", "Speciality", "Doctor NMC No.", "City", "Hospital"]
        + BS_MONTH_ABBR
        + ["Total Calls"]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, row in enumerate(rows, start=2):
        month_cells = [
            ", ".join(str(d) for d in mv) if mv else ""
            for mv in row["month_visits"]
        ]
        ws.append([
            row["msl"],
            row["doctor"].name,
            row["doctor"].specialization or "",
            row["doctor"].nmc_number,
            row["doctor"].area,
            " / ".join(row["hospitals"]),
            *month_cells,
            row["total_calls"],
        ])
        for cell in ws[i]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = [8, 22, 16, 14, 12, 22] + [10] * 12 + [10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    emp_name = employee.get_full_name() or employee.username
    response["Content-Disposition"] = (
        f'attachment; filename="yearly_report_{emp_name}_{year}.xlsx"'
    )
    wb.save(response)
    return response


# ── Monthly Target Report ────────────────────────────────────────────────────

def _build_target_rows(employee, year, month):
    """Rows shared by the Monthly Target report page and its Excel export."""
    relations = (
        DoctorEmployeeRelation.objects
        .filter(employee=employee, status=DoctorEmployeeRelation.Status.APPROVED)
        .select_related("doctor")
        .order_by("msl_number", "doctor__name")
    )

    visit_counts = defaultdict(int)
    for cov in DailyCoverage.objects.filter(
        created_by=employee,
        report_date__year=year,
        report_date__month=month,
    ).values("doctor_id"):
        visit_counts[cov["doctor_id"]] += 1

    rows = []
    for rel in relations:
        category = _doctor_category(rel.msl_number)
        target = VISIT_TARGETS[category]
        visits = visit_counts.get(rel.doctor_id, 0)

        if visits >= target:
            state = "green"
        elif visits > 0:
            state = "orange"
        else:
            state = "red"

        rows.append({
            "state":       state,
            "doctor_name": rel.doctor.name,
            "msl":         rel.msl_number or "-",
            "visits":      visits,
            "target":      target,
            "category":    CATEGORY_LABELS[category],
        })

    return rows


@login_required
@never_cache
def monthly_target_report(request):
    employee, all_employees, selected_employee_id, can_view_others = _get_employee(request)
    year, month, month_str = _parse_month(request.GET.get("month", ""))

    rows = _build_target_rows(employee, year, month)

    return render(request, "reports/monthly_target_report.html", {
        "month_str":           month_str,
        "year":                year,
        "month":               month,
        "month_name":          cal_module.month_name[month],
        "employee":            employee,
        "all_employees":       all_employees,
        "selected_employee_id": selected_employee_id,
        "can_view_others":     can_view_others,
        "rows":                rows,
    })


STATE_LABELS = {"green": "Met", "orange": "Partial", "red": "Not visited"}
STATE_FILLS = {"green": "C6EFCE", "orange": "FFEB9C", "red": "FFC7CE"}


@login_required
def monthly_target_report_excel(request):
    employee, _, _, _ = _get_employee(request)
    year, month, _ = _parse_month(request.GET.get("month", ""))
    rows = _build_target_rows(employee, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Target"

    header_fill = PatternFill("solid", fgColor="4A90C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(["Status", "Doctor's Name", "MSL No.", "Visits", "Target", "Class"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, row in enumerate(rows, start=2):
        ws.append([
            STATE_LABELS[row["state"]],
            row["doctor_name"],
            row["msl"],
            row["visits"],
            row["target"],
            row["category"],
        ])
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=STATE_FILLS[row["state"]])

    # Summary block
    ws.append([])
    counts = {state: sum(1 for r in rows if r["state"] == state) for state in STATE_LABELS}
    for state, label in STATE_LABELS.items():
        ws.append([label, counts[state]])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Total doctors", len(rows)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    for col_idx, width in enumerate([14, 26, 10, 8, 8, 12], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    emp_name = employee.get_full_name() or employee.username
    response["Content-Disposition"] = (
        f'attachment; filename="target_report_{emp_name}_{year}-{month:02d}.xlsx"'
    )
    wb.save(response)
    return response
